"""
Report Service
Business logic cho Report module.

Doc reference (2.4.3 – Báo cáo Tài chính):
1. Revenue Report: tổng doanh thu, theo khóa, theo lớp, ròng (sau hoàn tiền)
2. Expense Report: lương GV, breakdown full-time/part-time/native, thưởng KPI
3. Profit Report: Net profit = Revenue - Expense, profit margin
4. Payment Status: học viên nợ, GV chưa nhận lương, hoàn tiền đang xử lý
   → Router note chỉ có "debts" (học viên nợ). Doc bổ sung thêm GV chưa
     nhận lương + hoàn tiền đang xử lý, nhưng endpoint hiện tại chỉ có /debts.
     → Implement đúng theo router endpoint hiện có.

Export: Excel, CSV, PDF
"""
from sqlalchemy.orm import Session
from sqlalchemy import func, case, extract
from typing import List, Tuple, Optional
from uuid import UUID
from datetime import date, datetime, timezone
from decimal import Decimal
from fastapi import BackgroundTasks, HTTPException

from app.models.finance import (
    Invoice, InvoiceStatus,
    Payment, PaymentStatus,
    Refund, RefundStatus,
    ReportExportJob, ReportType, ExportJobStatus,
)
from app.models.academic import ClassEnrollment, Class, Course
from app.models.kpi import Salary, SalaryStatus, ContractType
from app.models.user import User

from app.schemas.finance.report import (
    RevenueReportResponse,
    CourseRevenueBreakdown,
    ExpensesReportResponse,
    ExpenseCategoryBreakdown,
    ProfitReportResponse,
    MonthlyTrend,
    DebtListResponse,
    ExportJobCreate,
    ExportJobResponse,
)


class ReportService:

    # -------------------------------------------------------------------
    # GET /reports/revenue
    # -------------------------------------------------------------------
    def get_revenue_report(
        self,
        db: Session,
        date_from: Optional[date],
        date_to: Optional[date],
        group_by_course: bool,
    ) -> RevenueReportResponse:
        """
        Aggregate payments SUCCESS trong khoảng thời gian.
        Doc: Tổng doanh thu, theo khóa, ròng (sau hoàn tiền).
        """
        query = db.query(Payment).filter(Payment.status == PaymentStatus.SUCCESS)

        if date_from:
            query = query.filter(func.date(Payment.paid_at) >= date_from)
        if date_to:
            query = query.filter(func.date(Payment.paid_at) <= date_to)

        # Aggregate
        result = query.with_entities(
            func.coalesce(func.sum(Payment.amount), 0).label("total_revenue"),
            func.count(Payment.id).label("total_invoices"),
        ).first()

        total_revenue = Decimal(str(result.total_revenue))
        total_invoices = result.total_invoices
        avg_payment_value = (
            (total_revenue / total_invoices).quantize(Decimal("0.01"))
            if total_invoices > 0
            else Decimal("0")
        )

        # Trừ hoàn tiền (doanh thu ròng) — theo doc
        refund_query = db.query(
            func.coalesce(func.sum(Refund.approved_amount), 0)
        ).filter(
            Refund.status.in_([RefundStatus.APPROVED, RefundStatus.PROCESSED]),
        )
        if date_from:
            refund_query = refund_query.filter(func.date(Refund.reviewed_at) >= date_from)
        if date_to:
            refund_query = refund_query.filter(func.date(Refund.reviewed_at) <= date_to)
        total_refunds = Decimal(str(refund_query.scalar() or 0))
        net_revenue = total_revenue - total_refunds

        # Breakdown by course nếu yêu cầu
        breakdown = None
        if group_by_course:
            breakdown = self._revenue_by_course(db, date_from, date_to)

        return RevenueReportResponse(
            date_from=date_from,
            date_to=date_to,
            total_revenue=net_revenue,
            total_invoices=total_invoices,
            avg_payment_value=avg_payment_value,
            breakdown_by_course=breakdown,
        )

    # -------------------------------------------------------------------
    # GET /reports/expenses
    # -------------------------------------------------------------------
    def get_expenses_report(
        self,
        db: Session,
        date_from: Optional[date],
        date_to: Optional[date],
        cost_type: Optional[str],
    ) -> ExpensesReportResponse:
        """
        Doc: Tổng lương GV, breakdown full-time/part-time/native, thưởng KPI.
        cost_type: SALARY | ALL (hiện tại chỉ hỗ trợ salary-based expenses)
        """
        salary_query = db.query(Salary)

        # Filter theo khoảng thời gian (period: YYYY-MM)
        if date_from:
            salary_query = salary_query.filter(Salary.period >= date_from.strftime("%Y-%m"))
        if date_to:
            salary_query = salary_query.filter(Salary.period <= date_to.strftime("%Y-%m"))

        # Breakdown theo contract type (doc: full-time vs part-time)
        breakdown_rows = (
            salary_query
            .with_entities(
                Salary.contract_type,
                func.coalesce(func.sum(Salary.net_salary), 0).label("total"),
            )
            .group_by(Salary.contract_type)
            .all()
        )

        breakdown = []
        total_expenses = Decimal("0")
        for row in breakdown_rows:
            amount = Decimal(str(row.total))
            total_expenses += amount
            breakdown.append(ExpenseCategoryBreakdown(
                category=row.contract_type.value if row.contract_type else "UNKNOWN",
                total=amount,
            ))

        # Thêm tổng thưởng KPI riêng biệt
        kpi_bonus_total = (
            salary_query
            .with_entities(func.coalesce(func.sum(Salary.kpi_bonus_calc), 0))
            .scalar()
        )
        kpi_bonus_decimal = Decimal(str(kpi_bonus_total))
        if kpi_bonus_decimal:
            total_expenses += kpi_bonus_decimal
            breakdown.append(ExpenseCategoryBreakdown(
                category="KPI_BONUS",
                total=kpi_bonus_decimal,
            ))

        return ExpensesReportResponse(
            date_from=date_from,
            date_to=date_to,
            cost_type=cost_type,
            total_expenses=total_expenses,
            breakdown_by_category=breakdown,
        )

    # -------------------------------------------------------------------
    # GET /reports/profit
    # -------------------------------------------------------------------
    def get_profit_report(
        self,
        db: Session,
        date_from: Optional[date],
        date_to: Optional[date],
    ) -> ProfitReportResponse:
        """
        Doc: Net profit = Revenue - Expense (lương + chi phí khác), Profit margin.
        """
        revenue_data = self.get_revenue_report(db, date_from, date_to, group_by_course=False)
        expenses_data = self.get_expenses_report(db, date_from, date_to, cost_type="ALL")

        total_revenue = revenue_data.total_revenue
        total_expenses = expenses_data.total_expenses
        profit = total_revenue - total_expenses
        profit_margin = (
            (profit / total_revenue * 100).quantize(Decimal("0.01"))
            if total_revenue > 0
            else Decimal("0")
        )

        # Monthly trends nếu có khoảng thời gian
        monthly_trends = None
        if date_from and date_to and date_from < date_to:
            monthly_trends = self._monthly_trends(db, date_from, date_to)

        return ProfitReportResponse(
            date_from=date_from,
            date_to=date_to,
            total_revenue=total_revenue,
            total_expenses=total_expenses,
            profit=profit,
            profit_margin=profit_margin,
            monthly_trends=monthly_trends,
        )

    # -------------------------------------------------------------------
    # GET /reports/debts
    # -------------------------------------------------------------------
    def get_debt_report(
        self, db: Session, page: int, limit: int
    ) -> Tuple[List[DebtListResponse], int]:
        """
        Doc (Payment Status): Học viên nợ (chưa thanh toán).
        Query Invoice status=PENDING mà due_date < now().
        """
        now = datetime.now(timezone.utc)

        query = (
            db.query(Invoice, User, Course)
            .join(User, User.id == Invoice.student_id)
            .outerjoin(ClassEnrollment, ClassEnrollment.id == Invoice.enrollment_id)
            .outerjoin(Class, Class.id == ClassEnrollment.class_id)
            .outerjoin(Course, Course.id == Class.course_id)
            .filter(
                Invoice.status == InvoiceStatus.PENDING,
                Invoice.due_date.isnot(None),
                Invoice.due_date < now,
                Invoice.deleted_at.is_(None),
            )
            .order_by(Invoice.final_amount.desc())
        )

        total = query.count()
        rows = query.offset((page - 1) * limit).limit(limit).all()

        items = []
        for invoice, user, course in rows:
            days_overdue = (now - invoice.due_date).days if invoice.due_date else 0
            items.append(DebtListResponse(
                invoice_id=invoice.id,
                student_id=user.id,
                student_name=f"{user.first_name} {user.last_name}",
                student_email=user.email,
                phone=user.phone,
                course_name=course.name if course else "—",
                debt_amount=invoice.final_amount,
                final_amount=invoice.final_amount,
                due_date=invoice.due_date,
                days_overdue=max(days_overdue, 0),
            ))

        return items, total

    # -------------------------------------------------------------------
    # POST /reports/export-jobs
    # -------------------------------------------------------------------
    def create_export_job(
        self,
        db: Session,
        payload: ExportJobCreate,
        created_by: UUID,
        bg_tasks: BackgroundTasks,
    ) -> ExportJobResponse:
        """
        Tạo job xuất báo cáo bất đồng bộ.
        Doc: Export format Excel, CSV, PDF.
        """
        job = ReportExportJob(
            report_type=payload.report_type,
            status=ExportJobStatus.PENDING,
            filters=payload.filters or {},
            created_by=created_by,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        # Enqueue background task
        bg_tasks.add_task(self._process_export, job.id)

        return ExportJobResponse.model_validate(job)

    # ===================================================================
    # Private helpers
    # ===================================================================

    def _revenue_by_course(
        self, db: Session, date_from: Optional[date], date_to: Optional[date]
    ) -> List[CourseRevenueBreakdown]:
        """Breakdown doanh thu theo khóa học."""
        query = (
            db.query(
                Course.id.label("course_id"),
                Course.name.label("course_name"),
                func.coalesce(func.sum(Payment.amount), 0).label("total_revenue"),
                func.count(Payment.id).label("total_invoices"),
            )
            .join(Invoice, Invoice.id == Payment.invoice_id)
            .join(ClassEnrollment, ClassEnrollment.id == Invoice.enrollment_id)
            .join(Class, Class.id == ClassEnrollment.class_id)
            .join(Course, Course.id == Class.course_id)
            .filter(Payment.status == PaymentStatus.SUCCESS)
        )

        if date_from:
            query = query.filter(func.date(Payment.paid_at) >= date_from)
        if date_to:
            query = query.filter(func.date(Payment.paid_at) <= date_to)

        rows = query.group_by(Course.id, Course.name).all()

        return [
            CourseRevenueBreakdown(
                course_id=row.course_id,
                course_name=row.course_name,
                total_revenue=Decimal(str(row.total_revenue)),
                total_invoices=row.total_invoices,
            )
            for row in rows
        ]

    def _monthly_trends(
        self, db: Session, date_from: date, date_to: date
    ) -> List[MonthlyTrend]:
        """Month-over-month revenue/expense/profit trends."""
        # Revenue by month
        rev_rows = (
            db.query(
                func.to_char(Payment.paid_at, "YYYY-MM").label("month"),
                func.coalesce(func.sum(Payment.amount), 0).label("revenue"),
            )
            .filter(
                Payment.status == PaymentStatus.SUCCESS,
                func.date(Payment.paid_at) >= date_from,
                func.date(Payment.paid_at) <= date_to,
            )
            .group_by(func.to_char(Payment.paid_at, "YYYY-MM"))
            .all()
        )
        revenue_map = {r.month: Decimal(str(r.revenue)) for r in rev_rows}

        # Expenses by month (salary period = YYYY-MM)
        exp_rows = (
            db.query(
                Salary.period.label("month"),
                func.coalesce(func.sum(Salary.net_salary), 0).label("expenses"),
            )
            .filter(
                Salary.period >= date_from.strftime("%Y-%m"),
                Salary.period <= date_to.strftime("%Y-%m"),
            )
            .group_by(Salary.period)
            .all()
        )
        expense_map = {e.month: Decimal(str(e.expenses)) for e in exp_rows}

        # Merge months
        all_months = sorted(set(list(revenue_map.keys()) + list(expense_map.keys())))
        trends = []
        for month in all_months:
            rev = revenue_map.get(month, Decimal("0"))
            exp = expense_map.get(month, Decimal("0"))
            trends.append(MonthlyTrend(
                month=month,
                revenue=rev,
                expenses=exp,
                profit=rev - exp,
            ))
        return trends

    def _process_export(self, job_id: UUID):
        """
        Background task: generate export file.
        Trong production sẽ dùng openpyxl/reportlab rồi upload S3.
        """
        from app.core.database import SessionLocal
        db = SessionLocal()
        try:
            job = db.query(ReportExportJob).filter(ReportExportJob.id == job_id).first()
            if not job:
                return

            job.status = ExportJobStatus.PROCESSING
            db.commit()

            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import os

            filters = job.filters or {}
            date_from_str = filters.get("date_from")
            date_to_str = filters.get("date_to")

            date_from = None
            if date_from_str:
                try:
                    date_from = date.fromisoformat(date_from_str[:10])
                except Exception:
                    pass
            date_to = None
            if date_to_str:
                try:
                    date_to = date.fromisoformat(date_to_str[:10])
                except Exception:
                    pass

            os.makedirs("media/exports", exist_ok=True)
            file_path = f"media/exports/{job.id}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Báo cáo"

            # Show grid lines
            ws.views.sheetView[0].showGridLines = True

            # Styling helpers
            font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
            font_section = Font(name="Segoe UI", size=13, bold=True, color="1F497D")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Segoe UI", size=11, bold=True)
            font_normal = Font(name="Segoe UI", size=11)
            font_meta = Font(name="Segoe UI", size=10, italic=True, color="595959")

            fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            fill_summary_label = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            if job.report_type == ReportType.REVENUE:
                rep = self.get_revenue_report(db, date_from, date_to, group_by_course=True)
                
                # Title
                ws.append(["BÁO CÁO DOANH THU"])
                ws.cell(1, 1).font = font_title
                ws.row_dimensions[1].height = 30
                
                # Metadata
                ws.append([f"Từ ngày: {date_from_str or 'Tất cả'}"])
                ws.cell(2, 1).font = font_meta
                ws.append([f"Đến ngày: {date_to_str or 'Tất cả'}"])
                ws.cell(3, 1).font = font_meta
                ws.append([]) # Row 4 empty
                
                # Key Summary Card
                ws.append(["Tổng doanh thu", rep.total_revenue])
                ws.append(["Số lượng hóa đơn", rep.total_invoices])
                ws.append(["Trung bình mỗi hóa đơn", rep.avg_payment_value])
                
                # Style key summary card
                for r in range(5, 8):
                    ws.cell(r, 1).font = font_bold
                    ws.cell(r, 1).fill = fill_summary_label
                    ws.cell(r, 1).border = thin_border
                    ws.cell(r, 2).font = font_bold
                    ws.cell(r, 2).border = thin_border
                    if r in (5, 7):
                        ws.cell(r, 2).number_format = '#,##0" đ"'
                        ws.cell(r, 2).alignment = align_right
                    else:
                        ws.cell(r, 2).number_format = '#,##0'
                        ws.cell(r, 2).alignment = align_right
                        
                ws.append([]) # Row 8 empty
                ws.append([]) # Row 9 empty
                
                # Course details header
                ws.append(["CHI TIẾT DOANH THU THEO KHÓA HỌC"])
                ws.cell(10, 1).font = font_section
                
                # Table headers
                headers = ["Khóa học", "Doanh thu", "Số hóa đơn"]
                ws.append(headers)
                ws.row_dimensions[11].height = 25
                for col_idx in range(1, 4):
                    cell = ws.cell(11, col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                # Table rows
                start_row = 12
                for idx, b in enumerate(rep.breakdown_by_course or []):
                    row_data = [b.course_name, b.total_revenue, b.total_invoices]
                    ws.append(row_data)
                    curr_row = start_row + idx
                    ws.row_dimensions[curr_row].height = 20
                    
                    # Col 1: Course Name
                    c1 = ws.cell(curr_row, 1)
                    c1.font = font_normal
                    c1.border = thin_border
                    c1.alignment = align_left
                    if idx % 2 == 1:
                        c1.fill = fill_zebra
                        
                    # Col 2: Revenue
                    c2 = ws.cell(curr_row, 2)
                    c2.font = font_normal
                    c2.border = thin_border
                    c2.number_format = '#,##0" đ"'
                    c2.alignment = align_right
                    if idx % 2 == 1:
                        c2.fill = fill_zebra
                        
                    # Col 3: Invoice count
                    c3 = ws.cell(curr_row, 3)
                    c3.font = font_normal
                    c3.border = thin_border
                    c3.number_format = '#,##0'
                    c3.alignment = align_center
                    if idx % 2 == 1:
                        c3.fill = fill_zebra

            elif job.report_type == ReportType.EXPENSES:
                rep = self.get_expenses_report(db, date_from, date_to, cost_type="ALL")
                
                # Title
                ws.append(["BÁO CÁO CHI PHÍ"])
                ws.cell(1, 1).font = font_title
                ws.row_dimensions[1].height = 30
                
                # Metadata
                ws.append([f"Từ ngày: {date_from_str or 'Tất cả'}"])
                ws.cell(2, 1).font = font_meta
                ws.append([f"Đến ngày: {date_to_str or 'Tất cả'}"])
                ws.cell(3, 1).font = font_meta
                ws.append([]) # Row 4 empty
                
                # Key Summary Card
                ws.append(["Tổng chi phí", rep.total_expenses])
                ws.cell(5, 1).font = font_bold
                ws.cell(5, 1).fill = fill_summary_label
                ws.cell(5, 1).border = thin_border
                ws.cell(5, 2).font = font_bold
                ws.cell(5, 2).border = thin_border
                ws.cell(5, 2).number_format = '#,##0" đ"'
                ws.cell(5, 2).alignment = align_right
                
                ws.append([]) # Row 6 empty
                ws.append([]) # Row 7 empty
                
                # Section header
                ws.append(["CHI TIẾT CHI PHÍ THEO DANH MỤC"])
                ws.cell(8, 1).font = font_section
                
                # Table headers
                headers = ["Danh mục", "Số tiền"]
                ws.append(headers)
                ws.row_dimensions[9].height = 25
                for col_idx in range(1, 3):
                    cell = ws.cell(9, col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                # Table rows
                start_row = 10
                category_labels = {
                    "FULL_TIME": "Lương cố định (Full-time)",
                    "PART_TIME": "Lương theo giờ (Part-time)",
                    "NATIVE": "Lương GV bản xứ",
                    "KPI_BONUS": "Thưởng KPI giáo viên",
                    "FACILITY": "Cơ sở vật chất",
                    "MARKETING": "Marketing",
                    "UTILITY": "Điện nước",
                    "OTHER": "Khác",
                }
                
                for idx, b in enumerate(rep.breakdown_by_category or []):
                    lbl = category_labels.get(b.category, b.category)
                    row_data = [lbl, b.total]
                    ws.append(row_data)
                    curr_row = start_row + idx
                    ws.row_dimensions[curr_row].height = 20
                    
                    # Col 1: Category
                    c1 = ws.cell(curr_row, 1)
                    c1.font = font_normal
                    c1.border = thin_border
                    c1.alignment = align_left
                    if idx % 2 == 1:
                        c1.fill = fill_zebra
                        
                    # Col 2: Amount
                    c2 = ws.cell(curr_row, 2)
                    c2.font = font_normal
                    c2.border = thin_border
                    c2.number_format = '#,##0" đ"'
                    c2.alignment = align_right
                    if idx % 2 == 1:
                        c2.fill = fill_zebra

            elif job.report_type == ReportType.PROFIT:
                rep = self.get_profit_report(db, date_from, date_to)
                
                # Title
                ws.append(["BÁO CÁO LỢI NHUẬN TỔNG HỢP"])
                ws.cell(1, 1).font = font_title
                ws.row_dimensions[1].height = 30
                
                # Metadata
                ws.append([f"Từ ngày: {date_from_str or 'Tất cả'}"])
                ws.cell(2, 1).font = font_meta
                ws.append([f"Đến ngày: {date_to_str or 'Tất cả'}"])
                ws.cell(3, 1).font = font_meta
                ws.append([]) # Row 4 empty
                
                # Summary card
                ws.append(["Doanh thu", rep.total_revenue])
                ws.append(["Chi phí", rep.total_expenses])
                ws.append(["Lợi nhuận ròng", rep.profit])
                ws.append(["Biên lợi nhuận", rep.profit_margin / 100])
                
                # Style summary card
                for r in range(5, 9):
                    ws.cell(r, 1).font = font_bold
                    ws.cell(r, 1).fill = fill_summary_label
                    ws.cell(r, 1).border = thin_border
                    ws.cell(r, 2).font = font_bold
                    ws.cell(r, 2).border = thin_border
                    if r in (5, 6, 7):
                        ws.cell(r, 2).number_format = '#,##0" đ"'
                        ws.cell(r, 2).alignment = align_right
                    else:
                        ws.cell(r, 2).number_format = '0.00%'
                        ws.cell(r, 2).alignment = align_right
                        
                ws.append([]) # Row 9 empty
                ws.append([]) # Row 10 empty
                
                # Section Header
                ws.append(["XU HƯỚNG THEO THÁNG"])
                ws.cell(11, 1).font = font_section
                
                # Table headers
                headers = ["Tháng", "Doanh thu", "Chi phí", "Lợi nhuận"]
                ws.append(headers)
                ws.row_dimensions[12].height = 25
                for col_idx in range(1, 5):
                    cell = ws.cell(12, col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                # Table rows
                start_row = 13
                for idx, t in enumerate(rep.monthly_trends or []):
                    row_data = [t.month, t.revenue, t.expenses, t.profit]
                    ws.append(row_data)
                    curr_row = start_row + idx
                    ws.row_dimensions[curr_row].height = 20
                    
                    # Col 1: Month
                    c1 = ws.cell(curr_row, 1)
                    c1.font = font_normal
                    c1.border = thin_border
                    c1.alignment = align_center
                    if idx % 2 == 1:
                        c1.fill = fill_zebra
                        
                    # Col 2-4: Revenue, Expenses, Profit
                    for col_idx in range(2, 5):
                        c = ws.cell(curr_row, col_idx)
                        c.font = font_normal
                        c.border = thin_border
                        c.number_format = '#,##0" đ"'
                        c.alignment = align_right
                        if idx % 2 == 1:
                            c.fill = fill_zebra

            elif job.report_type == ReportType.DEBTS:
                items, total = self.get_debt_report(db, page=1, limit=1000)
                
                # Title
                ws.append(["BÁO CÁO CÔNG NỢ HỌC VIÊN"])
                ws.cell(1, 1).font = font_title
                ws.row_dimensions[1].height = 30
                
                # Summary
                ws.append(["Tổng số học viên nợ phí", total])
                ws.cell(2, 1).font = font_bold
                ws.cell(2, 1).fill = fill_summary_label
                ws.cell(2, 1).border = thin_border
                ws.cell(2, 2).font = font_bold
                ws.cell(2, 2).border = thin_border
                ws.cell(2, 2).number_format = '#,##0'
                ws.cell(2, 2).alignment = align_right
                
                ws.append([]) # Row 3 empty
                
                # Table headers
                headers = ["Học viên", "SĐT", "Khóa học", "Số tiền nợ", "Hạn thanh toán", "Quá hạn"]
                ws.append(headers)
                ws.row_dimensions[4].height = 25
                for col_idx in range(1, 7):
                    cell = ws.cell(4, col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                # Table rows
                start_row = 5
                for idx, d in enumerate(items):
                    due_str = d.due_date.strftime("%Y-%m-%d") if d.due_date else "—"
                    row_data = [d.student_name, str(d.phone) if d.phone else "—", d.course_name or "—", d.debt_amount, due_str, f"{d.days_overdue} ngày"]
                    ws.append(row_data)
                    curr_row = start_row + idx
                    ws.row_dimensions[curr_row].height = 20
                    
                    # Col 1: Student Name
                    c1 = ws.cell(curr_row, 1)
                    c1.font = font_normal
                    c1.border = thin_border
                    c1.alignment = align_left
                    if idx % 2 == 1:
                        c1.fill = fill_zebra
                        
                    # Col 2: Phone (Save as string to prevent scientific notation)
                    c2 = ws.cell(curr_row, 2)
                    c2.font = font_normal
                    c2.border = thin_border
                    c2.number_format = '@' # Text format
                    c2.alignment = align_center
                    if idx % 2 == 1:
                        c2.fill = fill_zebra
                        
                    # Col 3: Course Name
                    c3 = ws.cell(curr_row, 3)
                    c3.font = font_normal
                    c3.border = thin_border
                    c3.alignment = align_left
                    if idx % 2 == 1:
                        c3.fill = fill_zebra
                        
                    # Col 4: Debt Amount
                    c4 = ws.cell(curr_row, 4)
                    c4.font = font_normal
                    c4.border = thin_border
                    c4.number_format = '#,##0" đ"'
                    c4.alignment = align_right
                    if idx % 2 == 1:
                        c4.fill = fill_zebra
                        
                    # Col 5: Due Date
                    c5 = ws.cell(curr_row, 5)
                    c5.font = font_normal
                    c5.border = thin_border
                    c5.alignment = align_center
                    if idx % 2 == 1:
                        c5.fill = fill_zebra
                        
                    # Col 6: Days Overdue
                    c6 = ws.cell(curr_row, 6)
                    c6.font = font_normal
                    c6.border = thin_border
                    c6.alignment = align_center
                    if idx % 2 == 1:
                        c6.fill = fill_zebra

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Ignore the title row when calculating column widths to avoid excessively wide columns
                    if cell.row == 1:
                        continue
                    val = str(cell.value or '')
                    # If it's a formatted number, add some estimated padding
                    if cell.number_format and ('đ' in cell.number_format or '%' in cell.number_format):
                        max_len = max(max_len, len(val) + 6)
                    else:
                        max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(file_path)

            job.status = ExportJobStatus.COMPLETED
            job.file_url = f"/media/exports/{job.id}.xlsx"
            job.completed_at = datetime.now(timezone.utc)
            db.commit()

        except Exception as e:
            job = db.query(ReportExportJob).filter(ReportExportJob.id == job_id).first()
            if job:
                job.status = ExportJobStatus.FAILED
                job.error_message = str(e)
                job.completed_at = datetime.now(timezone.utc)
                db.commit()
        finally:
            db.close()


report_service = ReportService()
